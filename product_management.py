import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

# ตัวแปรสำหรับเก็บที่อยู่ไฟล์ Excel
file_path = None
tree_data = []  # ตัวแปรเก็บข้อมูลทั้งหมดของ Treeview


# ฟังก์ชันสำหรับเพิ่มรายการสินค้าและคำนวณราคารวมทั้งหมด
def add_item():
    name = entry_name.get()
    try:
        price = float(entry_price.get())
        quantity = int(entry_quantity.get())
    except ValueError:
        return

    total_price = price * quantity
    tree.insert("", "end", values=(name, price, quantity, total_price))
    update_total()
    update_tree_data()
    if file_path:
        save_to_excel()

    entry_name.delete(0, tk.END)
    entry_price.delete(0, tk.END)
    entry_quantity.delete(0, tk.END)


# ฟังก์ชันสำหรับคำนวณราคารวมทั้งหมด
def update_total():
    total_price = 0
    total_items = 0
    for row in tree.get_children():
        values = tree.item(row)["values"]
        total_price += float(values[3])  # ราคารวม
        total_items += int(values[2])   # จำนวนสินค้า

    label_total.config(text=f"ราคารวมทั้งหมด: {total_price:.2f} บาท")
    label_total_items.config(text=f"จำนวนสินค้าทั้งหมด: {total_items} ชิ้น")



# ฟังก์ชันสำหรับบันทึกข้อมูลใน Treeview ไปยังไฟล์ Excel
def save_to_excel():
    if not file_path:
        save_file_as()
        return

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "รายการสินค้า"
        headers = ["ชื่อสินค้า", "ราคา/ชิ้น", "จำนวน", "ราคารวม"]
        for col_num, header in enumerate(headers, 1):
            ws[f"{get_column_letter(col_num)}1"] = header

    ws.delete_rows(2, ws.max_row)  # ลบข้อมูลเดิมก่อนเพิ่มใหม่

    for row in tree.get_children():
        values = tree.item(row)["values"]
        ws.append(values)

    wb.save(file_path)


# ฟังก์ชันสำหรับลบรายการสินค้า
def delete_item():
    selected_item = tree.selection()
    if selected_item:
        tree.delete(selected_item)
        update_tree_data()
        if file_path:
            save_to_excel()
        update_total()


# ฟังก์ชันสำหรับโหลดข้อมูลจาก Excel
def load_from_excel():
    if file_path and os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        for row in tree.get_children():
            tree.delete(row)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            tree.insert("", "end", values=row)
        update_total()
        update_tree_data()


# ฟังก์ชันสำหรับเลือกที่อยู่ในการบันทึกไฟล์
def save_file_as():
    global file_path
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        save_to_excel()


# ฟังก์ชันสำหรับเปิดไฟล์
def open_file():
    global file_path
    file_path = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        load_from_excel()


# ฟังก์ชันสำหรับค้นหารายการสินค้า
def search_item(event=None):
    query = entry_search.get().lower()
    for item in tree.get_children():
        tree.delete(item)

    if query:
        for row in tree_data:
            if query in str(row[0]).lower():
                tree.insert("", "end", values=row)
    else:
        for row in tree_data:
            tree.insert("", "end", values=row)


# ฟังก์ชันสำหรับอัปเดตข้อมูลใน tree_data
def update_tree_data():
    global tree_data
    tree_data = []
    for row in tree.get_children():
        tree_data.append(tree.item(row)["values"])


# ฟังก์ชันสำหรับการแก้ไขข้อมูลใน Treeview
is_edit_window_open = False

def edit_item(event):
    global is_edit_window_open
    if is_edit_window_open:
        return  # หากมีหน้าต่างเปิดอยู่แล้ว ให้ยกเลิกการทำงาน

    selected_item = tree.selection()
    if not selected_item:
        return

    is_edit_window_open = True  # ตั้งค่าสถานะเป็น True เมื่อเปิดหน้าต่าง
    item = tree.item(selected_item)
    values = item["values"]

    edit_window = tk.Toplevel(root)
    edit_window.title("แก้ไขข้อมูลสินค้า")

    tk.Label(edit_window, text="ชื่อสินค้า:").grid(row=0, column=0, padx=5, pady=5)
    entry_edit_name = tk.Entry(edit_window)
    entry_edit_name.grid(row=0, column=1, padx=5, pady=5)
    entry_edit_name.insert(0, values[0])

    tk.Label(edit_window, text="ราคาต่อชิ้น:").grid(row=1, column=0, padx=5, pady=5)
    entry_edit_price = tk.Entry(edit_window)
    entry_edit_price.grid(row=1, column=1, padx=5, pady=5)
    entry_edit_price.insert(0, values[1])

    tk.Label(edit_window, text="จำนวน:").grid(row=2, column=0, padx=5, pady=5)
    entry_edit_quantity = tk.Entry(edit_window)
    entry_edit_quantity.grid(row=2, column=1, padx=5, pady=5)
    entry_edit_quantity.insert(0, values[2])

    def save_edit():
        try:
            new_name = entry_edit_name.get()
            new_price = float(entry_edit_price.get())
            new_quantity = int(entry_edit_quantity.get())
            new_total = new_price * new_quantity
        except ValueError:
            messagebox.showerror("ข้อผิดพลาด", "กรุณากรอกข้อมูลที่ถูกต้อง")
            return

        tree.item(selected_item, values=(new_name, new_price, new_quantity, new_total))
        update_total()
        update_tree_data()
        if file_path:
            save_to_excel()
        close_edit_window()  # ปิดหน้าต่างแก้ไขเมื่อบันทึกเสร็จ

    def close_edit_window():
        global is_edit_window_open
        is_edit_window_open = False  # ตั้งสถานะเป็น False เมื่อปิดหน้าต่าง
        edit_window.destroy()

    edit_window.protocol("WM_DELETE_WINDOW", close_edit_window)  # จัดการเมื่อกดปุ่มปิดหน้าต่าง
    tk.Button(edit_window, text="บันทึก", command=save_edit).grid(row=3, column=0, columnspan=2, pady=5)



# สร้างหน้าต่างหลัก
root = tk.Tk()
root.title("โปรแกรมเพิ่มรายการสินค้า")

# สร้างกรอบสำหรับการกรอกข้อมูลและปุ่ม
frame_input = tk.Frame(root)
frame_input.pack(fill='x', padx=10, pady=5)

label_name = tk.Label(frame_input, text="ชื่อสินค้า:")
label_name.grid(row=0, column=0, padx=5, pady=5)
entry_name = tk.Entry(frame_input)
entry_name.grid(row=0, column=1, padx=5, pady=5)

label_price = tk.Label(frame_input, text="ราคาต่อชิ้น:")
label_price.grid(row=1, column=0, padx=5, pady=5)
entry_price = tk.Entry(frame_input)
entry_price.grid(row=1, column=1, padx=5, pady=5)

label_quantity = tk.Label(frame_input, text="จำนวน:")
label_quantity.grid(row=2, column=0, padx=5, pady=5)
entry_quantity = tk.Entry(frame_input)
entry_quantity.grid(row=2, column=1, padx=5, pady=5)

button_add = tk.Button(frame_input, text="เพิ่มสินค้า", command=add_item)
button_add.grid(row=0, column=2, padx=5)

button_delete = tk.Button(frame_input, text="ลบสินค้า", command=delete_item)
button_delete.grid(row=1, column=2, padx=5)

# ช่องค้นหา
frame_search = tk.Frame(root)
frame_search.pack(fill='x', padx=10, pady=5)

label_search = tk.Label(frame_search, text="ค้นหา:")
label_search.grid(row=0, column=0, padx=18, pady=5)
entry_search = tk.Entry(frame_search)
entry_search.grid(row=0, column=1, padx=2, pady=5)
entry_search.bind("<KeyRelease>", search_item)

# สร้างตาราง
frame_table = tk.Frame(root)
frame_table.pack(fill='x', padx=10, pady=5)

tree = ttk.Treeview(frame_table, columns=("name", "price", "quantity", "total"), show="headings")
tree.heading("name", text="ชื่อสินค้า")
tree.heading("price", text="ราคา/ชิ้น")
tree.heading("quantity", text="จำนวน")
tree.heading("total", text="ราคารวม")
tree.pack(fill="both", expand=True)

# ผูกเหตุการณ์ดับเบิลคลิกกับ Treeview
tree.bind("<Double-1>", edit_item)

# สร้างกรอบสำหรับแสดงจำนวนสินค้ารวมและราคารวมในบรรทัดเดียวกัน
frame_summary = tk.Frame(root)
frame_summary.pack(fill='x', padx=10, pady=5)

# แสดงจำนวนสินค้ารวม
label_total_items = tk.Label(frame_summary, text="จำนวนสินค้าทั้งหมด: 0 ชิ้น")
label_total_items.pack(side="left", padx=5)

# แสดงราคารวม
label_total = tk.Label(frame_summary, text="ราคารวมทั้งหมด: 0.00 บาท")
label_total.pack(side="left", padx=5)

menu = tk.Menu(root)
root.config(menu=menu)
file_menu = tk.Menu(menu, tearoff=0)
menu.add_cascade(label="ไฟล์", menu=file_menu)
file_menu.add_command(label="เปิดไฟล์", command=open_file)
file_menu.add_command(label="บันทึกไฟล์", command=save_file_as)
file_menu.add_command(label="ออก", command=root.quit)

root.mainloop()
